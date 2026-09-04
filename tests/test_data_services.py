import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from data_services import (
    DataFormatError,
    PersonalSchedule,
    _parse_personal_schedule_rows,
    export_schedule_workbook,
    load_personal_excel,
    personal_course_color,
    personal_schedule_from_dict,
)
from models import Course, TimeSlot


class PersonalScheduleTests(unittest.TestCase):
    def test_json_schedule_detects_week_aware_conflicts(self):
        schedule = personal_schedule_from_dict({
            "term": "2026-2027-1",
            "slots": [{"weekday": 2, "sections": [1, 2], "weeks": [1, 3]}],
        })
        conflict = Course("A", "Conflict", "T", (TimeSlot(2, frozenset({2}), frozenset({3}), ""),), {}, "", "#fff")
        alternate = Course("B", "Alternate", "T", (TimeSlot(2, frozenset({2}), frozenset({2}), ""),), {}, "", "#fff")
        self.assertTrue(schedule.conflicts(conflict))
        self.assertFalse(schedule.conflicts(alternate))

    def test_json_slot_validation_is_consistent(self):
        invalid_slots = (
            {"sections": [1], "weeks": [1]},
            {"weekday": 1, "weeks": [1]},
            {"weekday": "星期一", "sections": [1], "weeks": [1]},
            {"weekday": 1, "sections": ["第一节"], "weeks": [1]},
            {"weekday": 8, "sections": [1], "weeks": [1]},
            {"weekday": 1, "sections": [12], "weeks": [1]},
            {"weekday": 1, "sections": [1], "weeks": []},
        )
        for slot in invalid_slots:
            with self.subTest(slot=slot), self.assertRaises(DataFormatError):
                personal_schedule_from_dict({"slots": [slot]})

    def test_json_slots_must_be_lists(self):
        with self.assertRaises(DataFormatError):
            personal_schedule_from_dict({"slots": {"weekday": 1, "sections": [1]}})
        with self.assertRaises(DataFormatError):
            personal_schedule_from_dict({
                "slots": [],
                "courses": [{"name": "错误课程", "slots": {"weekday": 1, "sections": [1]}}],
            })

    def test_course_slots_use_the_same_json_validation(self):
        invalid_course_slots = (
            {"sections": [1], "weeks": [1]},
            {"weekday": 1, "weeks": [1]},
            {"weekday": "星期一", "sections": [1], "weeks": [1]},
            {"weekday": 1, "sections": ["第一节"], "weeks": [1]},
            {"weekday": 8, "sections": [1], "weeks": [1]},
            {"weekday": 1, "sections": [12], "weeks": [1]},
        )
        for slot in invalid_course_slots:
            with self.subTest(slot=slot), self.assertRaises(DataFormatError):
                personal_schedule_from_dict({
                    "slots": [],
                    "courses": [{"name": "错误课程", "slots": [slot]}],
                })

    def test_valid_json_slots_support_sections_and_periods(self):
        schedule = personal_schedule_from_dict({
            "term": "2026-2027-1",
            "slots": [{"weekday": "2", "sections": ["1", 2], "weeks": [1, "3"]}],
            "courses": [{
                "name": "正常课程",
                "teacher": "测试教师",
                "slots": [{"weekday": 3, "periods": [10, 11], "weeks": [1, 2]}],
            }],
        })
        self.assertEqual(schedule.slots[0].periods, frozenset({1, 2}))
        self.assertEqual(schedule.slots[0].weeks, frozenset({1, 3}))
        self.assertEqual(schedule.courses[0].slots[0].periods, frozenset({10, 11}))

    def test_schedule_serialization_uses_public_schema(self):
        schedule = PersonalSchedule("2026-2027-1", (TimeSlot(1, frozenset({1, 2}), frozenset({1}), ""),))
        self.assertEqual(schedule.to_dict()["slots"][0]["sections"], [1, 2])

    def test_exported_hnu_workbook_preserves_course_details_and_exact_weeks(self):
        fixture = Path(__file__).with_name("_temporary_personal_schedule.xlsx")
        self.addCleanup(fixture.unlink, missing_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["学年学期：2026-2027-1"])
        sheet.append(["节次", "星期一", "星期二"])
        sheet.append(["01、02小节", "普通物理AⅡ\n孙辰教授;1-2,4周综合教学楼(综414)\nnull", ""])
        workbook.save(fixture)
        schedule = load_personal_excel(fixture)
        physics = next(course for course in schedule.courses if course.name == "普通物理AⅡ")
        self.assertEqual(physics.teacher, "孙辰教授")
        self.assertEqual(physics.classroom, "综合教学楼(综414)")
        self.assertNotIn(3, physics.slots[0].weeks)
        self.assertIn(4, physics.slots[0].weeks)

    def test_html_excel_export_is_detected_by_content(self):
        document = """<html><body><table>
        <tr><td>学年学期：2026-2027-1</td></tr>
        <tr><th>节次</th><th>星期一</th><th>星期二</th></tr>
        <tr><td>01,02小节</td><td>软件工程导论<br>肖雄仁;1-8周;中楼(中210)</td><td></td></tr>
        </table></body></html>"""
        path = Path(__file__).with_name("_temporary_html_export.xls")
        self.addCleanup(path.unlink, missing_ok=True)
        path.write_text(document, encoding="utf-8")
        schedule = load_personal_excel(path)
        self.assertEqual(schedule.diagnostics["format"], "html")
        self.assertEqual(schedule.courses[0].name, "软件工程导论")
        self.assertEqual(schedule.courses[0].classroom, "中楼(中210)")
        self.assertEqual(schedule.courses[0].slots[0].weeks, frozenset(range(1, 9)))

    def test_login_page_is_reported_instead_of_openpyxl_zip_error(self):
        path = Path(__file__).with_name("_temporary_login_page.xlsx")
        self.addCleanup(path.unlink, missing_ok=True)
        path.write_text("<html><body>请登录</body></html>", encoding="utf-8")
        with self.assertRaisesRegex(DataFormatError, "网页而不是课表文件"):
            load_personal_excel(path)

    def test_current_hnu_xls_layout_parses_periods_teacher_location_and_weeks(self):
        rows = [
            ["学年学期：2026-2027-1", "", ""],
            ["", "星期日", "星期一", "星期二", "星期三", "星期四", "星期五", "星期六"],
            [
                "第二大节\n03、04\n10:00~11:40",
                "",
                "",
                "",
                "",
                "",
                "软件工程导论\n肖雄仁副教授;4,8,12,16周中楼(中210)\nnull\n",
                "",
            ],
        ]
        schedule = _parse_personal_schedule_rows(rows, Path("current-hnu.xls"), "xls")
        course = schedule.courses[0]
        self.assertEqual(course.teacher, "肖雄仁副教授")
        self.assertEqual(course.classroom, "中楼(中210)")
        self.assertEqual(course.slots[0].periods, frozenset({3, 4}))
        self.assertEqual(course.slots[0].weeks, frozenset({4, 8, 12, 16}))
        self.assertEqual(course.color, personal_course_color("软件工程导论"))

    def test_remark_courses_are_preserved_without_timetable_slots(self):
        rows = [
            ["学年学期：2026-2027-1", ""],
            ["节次", "星期一"],
            ["01、02小节", "高等数学AⅠ\n严子龙教授;1-16周综合楼(综503)"],
            ["", "备注:军事技能 郭兰 5-6周;材料化学基础实验Ⅰ 范长岭,何月德 0-4,7-17周;"],
        ]

        schedule = _parse_personal_schedule_rows(rows, Path("remarks.xls"), "xls")
        remarks = [course for course in schedule.courses if not course.slots]
        self.assertEqual([course.name for course in remarks], ["军事技能", "材料化学基础实验Ⅰ"])
        self.assertEqual(remarks[1].teacher, "范长岭,何月德")
        self.assertEqual(schedule.diagnostics["remark_courses"], 2)

    def test_removing_imported_course_releases_all_of_its_slots(self):
        target = Course("A", "替换课程", "新老师", (TimeSlot(2, frozenset({3, 4}), frozenset({4}), ""),), {}, "", "#fff")
        imported = personal_schedule_from_dict({
            "slots": [{"weekday": 2, "sections": [3, 4], "weeks": [4]}],
            "courses": [{
                "name": "软件工程导论",
                "teacher": "原老师",
                "classroom": "原教室",
                "slots": [{"weekday": 2, "sections": [3, 4], "weeks": [4]}],
            }],
        })
        self.assertTrue(imported.conflicts(target))
        updated = imported.without_course("软件工程导论")
        self.assertFalse(updated.conflicts(target))
        self.assertFalse(updated.courses)

    def test_schedule_export_uses_importable_personal_timetable_layout(self):
        path = Path(__file__).with_name("_temporary_exported_schedule.xlsx")
        self.addCleanup(path.unlink, missing_ok=True)
        selected = Course(
            "ZH128XK24",
            "编译原理与技术",
            "杨金民",
            (TimeSlot(2, frozenset({3, 4}), frozenset(range(1, 17)), ""),),
            {},
            "",
            "#fff",
            classroom="综合楼(综310)",
        )
        personal = PersonalSchedule(
            "2026-2027-1",
            (),
            courses=(
                personal_schedule_from_dict({
                    "slots": [],
                    "courses": [{"name": "军事技能", "teacher": "郭兰", "details": "军事技能 郭兰 5-6周", "slots": []}],
                }).courses[0],
            ),
        )

        self.assertEqual(export_schedule_workbook(path, [selected], personal), 2)
        workbook = load_workbook(path)
        sheet = workbook.active
        self.assertEqual([sheet.cell(3, column).value for column in range(1, 9)], ["节次", "星期日", "星期一", "星期二", "星期三", "星期四", "星期五", "星期六"])
        self.assertIn("编译原理与技术", sheet.cell(5, 4).value)
        self.assertIn("备注:军事技能 郭兰 5-6周", sheet.cell(9, 2).value)

        imported = load_personal_excel(path)
        self.assertIn("编译原理与技术", [course.name for course in imported.courses])
        self.assertIn("军事技能", [course.name for course in imported.courses])


if __name__ == "__main__":
    unittest.main()
