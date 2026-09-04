"""Import, export, and persistence services for user-owned schedule data."""

from __future__ import annotations

import json
import hashlib
import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import date
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from models import Course, TimeSlot
from parser_utils import DEFAULT_WEEKS, parse_weeks


class DataFormatError(ValueError):
    pass


PERSONAL_COURSE_PALETTE = (
    "#F9C5D1", "#FDD9B5", "#FFF0B3", "#CDECCF",
    "#BDE8E9", "#C7DBF8", "#D8CCF2", "#F2CCE2",
)


def personal_course_color(name: str) -> str:
    digest = hashlib.blake2s(name.strip().encode("utf-8"), digest_size=4).digest()
    return PERSONAL_COURSE_PALETTE[int.from_bytes(digest, "big") % len(PERSONAL_COURSE_PALETTE)]


@dataclass(slots=True)
class PersonalCourse:
    name: str
    teacher: str
    classroom: str
    details: str
    slots: tuple[TimeSlot, ...]
    color: str = "#CDECCF"


@dataclass(slots=True)
class PersonalSchedule:
    term: str = ""
    slots: tuple[TimeSlot, ...] = ()
    source: str = ""
    diagnostics: dict[str, int | str] = field(default_factory=dict)
    courses: tuple[PersonalCourse, ...] = ()

    def conflicts(self, course: Course) -> bool:
        return any(personal.overlaps(candidate) for personal in self.slots for candidate in course.slots)

    def without_course(self, name: str) -> "PersonalSchedule":
        removed = [course for course in self.courses if course.name == name]
        if not removed:
            return self
        remaining = tuple(course for course in self.courses if course.name != name)

        # Preserve occupied slots that came from schedule formats without course metadata.
        owned_counts = Counter(slot for course in self.courses for slot in course.slots)
        anonymous_slots: list[TimeSlot] = []
        for slot in self.slots:
            if owned_counts[slot]:
                owned_counts[slot] -= 1
            else:
                anonymous_slots.append(slot)
        remaining_slots = tuple(anonymous_slots) + tuple(slot for course in remaining for slot in course.slots)
        diagnostics = dict(self.diagnostics)
        diagnostics["parsed_slots"] = len(remaining_slots)
        diagnostics["parsed_courses"] = len(remaining)
        return PersonalSchedule(self.term, remaining_slots, self.source, diagnostics, remaining)

    def to_dict(self) -> dict[str, Any]:
        return {
            "term": self.term,
            "source": self.source,
            "slots": [
                {"weekday": slot.weekday, "sections": sorted(slot.periods), "weeks": sorted(slot.weeks)}
                for slot in self.slots
            ],
            "diagnostics": self.diagnostics,
            "courses": [
                {
                    "name": course.name,
                    "teacher": course.teacher,
                    "classroom": course.classroom,
                    "details": course.details,
                    "color": course.color,
                    "slots": [
                        {"weekday": slot.weekday, "sections": sorted(slot.periods), "weeks": sorted(slot.weeks)}
                        for slot in course.slots
                    ],
                }
                for course in self.courses
            ],
        }


def _parse_plain_weeks(text: str) -> frozenset[int]:
    cleaned = text.replace("周", "").strip()
    return parse_weeks(f"({cleaned})周") if cleaned else DEFAULT_WEEKS


def _periods_from_label(value: object) -> frozenset[int]:
    text = str(value or "")
    match = re.search(
        r"^\s*((?:0[1-9]|1[01])(?:[、,，\s]+(?:0[1-9]|1[01]))*)\s*(?:小节)?\s*$",
        text,
        re.MULTILINE,
    )
    return frozenset(int(number) for number in re.findall(r"0[1-9]|1[0-4]", match.group(1))) if match else frozenset()


def _course_details(detail_line: str) -> tuple[str, str, list[str]]:
    teacher, separator, remainder = detail_line.partition(";")
    if not separator:
        teacher, separator, remainder = detail_line.partition("；")
    week_matches = re.findall(r"([0-9][0-9,，、\-~至\s]*周)", detail_line)
    classroom = "未知"
    if week_matches:
        week_end = detail_line.find(week_matches[-1]) + len(week_matches[-1])
        classroom = detail_line[week_end:].strip(" ;；") or "未知"
    elif separator:
        fields = [field.strip() for field in re.split(r"[;；]", remainder) if field.strip()]
        classroom = fields[-1] if fields else "未知"
    return teacher.strip() or "未知", classroom, week_matches


def _parse_remark_courses(rows: list[list[object]]) -> list[PersonalCourse]:
    courses: list[PersonalCourse] = []
    seen: set[tuple[str, str]] = set()
    for row in rows:
        for value in row:
            text = str(value or "").strip()
            match = re.search(r"(?:^|\n)\s*备注\s*[:：]\s*(.+)", text, re.DOTALL)
            if not match:
                continue
            for raw_entry in re.split(r"[;；]", match.group(1)):
                entry = raw_entry.strip()
                if not entry:
                    continue
                week_match = re.search(r"\s+[0-9][0-9,，、\-~至\s]*周\s*$", entry)
                identity = entry[:week_match.start()].strip() if week_match else entry
                parts = identity.rsplit(None, 1)
                name = parts[0].strip()
                teacher = parts[1].strip() if len(parts) == 2 else "未知"
                key = (name, teacher)
                if not name or key in seen:
                    continue
                seen.add(key)
                courses.append(PersonalCourse(name, teacher, "不排入课表", entry, (), personal_course_color(name)))
    return courses


class _HtmlTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.tables: list[list[list[str]]] = []
        self._table: list[list[str]] | None = None
        self._row: list[str] | None = None
        self._cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.lower()
        if tag == "table":
            self._table = []
        elif tag == "tr" and self._table is not None:
            self._row = []
        elif tag in {"td", "th"} and self._row is not None:
            self._cell = []
        elif tag == "br" and self._cell is not None:
            self._cell.append("\n")

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in {"td", "th"} and self._cell is not None and self._row is not None:
            self._row.append("".join(self._cell).strip())
            self._cell = None
        elif tag == "tr" and self._row is not None and self._table is not None:
            self._table.append(self._row)
            self._row = None
        elif tag == "table" and self._table is not None:
            self.tables.append(self._table)
            self._table = None


def _decode_web_document(content: bytes) -> str:
    for encoding in ("utf-8-sig", "gb18030", "utf-16"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _read_personal_schedule_rows(source: Path) -> tuple[list[list[object]], str]:
    try:
        content = source.read_bytes()
    except OSError as exc:
        raise DataFormatError(f"无法读取个人课表文件：{exc}") from exc
    if not content:
        raise DataFormatError("教务系统返回了空文件，请刷新课表页面后重新读取。")

    try:
        if content.startswith(b"PK\x03\x04"):
            workbook = load_workbook(source, data_only=True, read_only=True)
            try:
                rows = [[cell.value for cell in row] for row in workbook.active.iter_rows()]
            finally:
                workbook.close()
            return rows, "xlsx"
        if content.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
            import xlrd

            workbook = xlrd.open_workbook(source)
            try:
                sheet = workbook.sheet_by_index(0)
                rows = [sheet.row_values(index) for index in range(sheet.nrows)]
            finally:
                workbook.release_resources()
            return rows, "xls"
    except Exception as exc:
        raise DataFormatError(f"个人课表文件已识别，但内容损坏或不完整：{exc}") from exc

    preview = content[:4096].lstrip().lower()
    if preview.startswith((b"<!doctype html", b"<html", b"<table")) or b"<table" in preview:
        parser = _HtmlTableParser()
        parser.feed(_decode_web_document(content))
        candidates = [
            table for table in parser.tables
            if any("星期" in str(cell) for row in table for cell in row)
        ]
        if candidates:
            return max(candidates, key=len), "html"
        raise DataFormatError(
            "教务系统返回的是网页而不是课表文件。请确认登录仍有效、当前已打开“我的课表”，然后重试。"
        )
    raise DataFormatError(
        "教务系统返回了不受支持的文件格式（不是 XLSX、XLS 或可解析的网页课表）。"
    )


def _parse_personal_schedule_rows(rows: list[list[object]], source: Path, detected_format: str) -> PersonalSchedule:
    header_row = None
    day_columns: dict[int, int] = {}
    term = ""
    for row_index, row in enumerate(rows):
        for column_index, value in enumerate(row):
            text = str(value or "").strip()
            term_match = re.search(r"学年学期[：:]\s*([^\s]+)", text)
            if term_match:
                term = term_match.group(1)
            day_match = re.fullmatch(r"星期([一二三四五六日天])", text)
            if day_match:
                header_row = row_index
                day_columns[column_index] = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "日": 7, "天": 7}[day_match.group(1)]
    if header_row is None or not day_columns:
        raise DataFormatError("未找到“星期一”至“星期日”表头，这不是受支持的个人课表。")

    slots: list[TimeSlot] = []
    courses: list[PersonalCourse] = []
    occupied_cells = 0
    for row in rows[header_row + 1:]:
        periods = _periods_from_label(row[0] if row else None)
        if not periods:
            continue
        for column, weekday in day_columns.items():
            text = str(row[column] if column < len(row) else "").strip()
            if not text:
                continue
            occupied_cells += 1
            blocks = [block.strip() for block in re.split(r"\n\s*\n+", text) if block.strip()]
            parsed_block = False
            for block in blocks:
                lines = [line.strip() for line in block.splitlines() if line.strip() and line.strip().lower() != "null"]
                detail_index = next((index for index, line in enumerate(lines) if ";" in line and "周" in line), None)
                if detail_index is None:
                    continue
                name = lines[0]
                teacher, classroom, week_matches = _course_details(lines[detail_index])
                weeks = frozenset().union(*(_parse_plain_weeks(value) for value in week_matches)) if week_matches else DEFAULT_WEEKS
                slot = TimeSlot(weekday, periods, frozenset(weeks), "、".join(week_matches))
                slots.append(slot)
                courses.append(PersonalCourse(name, teacher, classroom, block, (slot,), personal_course_color(name)))
                parsed_block = True
            if not parsed_block:
                week_matches = re.findall(r"([0-9][0-9,，、\-~至\s]*周)", text)
                weeks = frozenset().union(*(_parse_plain_weeks(value) for value in week_matches)) if week_matches else DEFAULT_WEEKS
                slots.append(TimeSlot(weekday, periods, frozenset(weeks), "、".join(week_matches)))
    remark_courses = _parse_remark_courses(rows)
    courses.extend(remark_courses)
    if not slots and not courses:
        raise DataFormatError("课表中没有解析到任何已占用时间。")
    return PersonalSchedule(term, tuple(slots), str(source), {"format": detected_format, "weekday_columns": len(day_columns), "occupied_cells": occupied_cells, "parsed_slots": len(slots), "parsed_courses": len(courses), "remark_courses": len(remark_courses)}, tuple(courses))


def load_personal_excel(path: str | Path) -> PersonalSchedule:
    source = Path(path)
    rows, detected_format = _read_personal_schedule_rows(source)
    return _parse_personal_schedule_rows(rows, source, detected_format)


def _json_integer(value: object) -> int:
    if isinstance(value, bool):
        raise ValueError
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str) and re.fullmatch(r"[+-]?\d+", value.strip()):
        return int(value)
    raise ValueError


def _parse_slot_dict(item: object, context: str) -> TimeSlot:
    if not isinstance(item, dict):
        raise DataFormatError(f"{context}格式无效，时间段必须是 JSON 对象。")
    try:
        weekday = _json_integer(item["weekday"])
        if "sections" in item:
            raw_periods = item["sections"]
        elif "periods" in item:
            raw_periods = item["periods"]
        else:
            raise KeyError("sections")
        raw_weeks = item.get("weeks", list(range(1, 19)))
        if not isinstance(raw_periods, list) or not isinstance(raw_weeks, list):
            raise TypeError
        periods = frozenset(_json_integer(value) for value in raw_periods)
        weeks = frozenset(_json_integer(value) for value in raw_weeks)
    except (KeyError, TypeError, ValueError) as exc:
        raise DataFormatError(f"{context}格式无效。") from exc
    if not 1 <= weekday <= 7:
        raise DataFormatError(f"{context}的 weekday 必须在 1 至 7 之间。")
    if not periods or not all(1 <= value <= 11 for value in periods):
        raise DataFormatError(f"{context}的节次必须在 1 至 11 之间。")
    if not weeks or not all(1 <= value <= 30 for value in weeks):
        raise DataFormatError(f"{context}的周次必须在 1 至 30 之间。")
    try:
        return TimeSlot(weekday, periods, weeks, str(item.get("raw_weeks") or ""))
    except (TypeError, ValueError) as exc:
        raise DataFormatError(f"{context}格式无效。") from exc


def personal_schedule_from_dict(payload: object, source: str = "JSON") -> PersonalSchedule:
    if not isinstance(payload, dict) or not isinstance(payload.get("slots"), list):
        raise DataFormatError("JSON 必须包含 slots 数组。")
    raw_courses = payload.get("courses", [])
    if not isinstance(raw_courses, list):
        raise DataFormatError("JSON 中的 courses 必须是数组。")
    diagnostics = payload.get("diagnostics", {})
    if not isinstance(diagnostics, dict):
        raise DataFormatError("JSON 中的 diagnostics 必须是对象。")
    slots = [
        _parse_slot_dict(item, f"第 {index} 个时间段")
        for index, item in enumerate(payload["slots"], 1)
    ]
    courses: list[PersonalCourse] = []
    for course_index, item in enumerate(raw_courses, 1):
        if not isinstance(item, dict):
            raise DataFormatError(f"第 {course_index} 门课程格式无效，课程必须是 JSON 对象。")
        raw_course_slots = item.get("slots", [])
        if not isinstance(raw_course_slots, list):
            raise DataFormatError(f"第 {course_index} 门课程的 slots 必须是数组。")
        course_slots = [
            _parse_slot_dict(raw_slot, f"第 {course_index} 门课程的第 {slot_index} 个时间段")
            for slot_index, raw_slot in enumerate(raw_course_slots, 1)
        ]
        name = str(item.get("name", "已有课程"))
        courses.append(PersonalCourse(name, str(item.get("teacher", "未知")), str(item.get("classroom", "未知")), str(item.get("details", "")), tuple(course_slots), str(item.get("color") or personal_course_color(name))))
    return PersonalSchedule(str(payload.get("term", "")), tuple(slots), source, dict(diagnostics), tuple(courses))


def load_personal_json(path: str | Path) -> PersonalSchedule:
    source = Path(path)
    try:
        payload = json.loads(source.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise DataFormatError(f"无法读取个人课表 JSON：{exc}") from exc
    return personal_schedule_from_dict(payload, str(source))


def save_personal_json(schedule: PersonalSchedule, path: str | Path) -> None:
    Path(path).write_text(json.dumps(schedule.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_weeks(weeks: frozenset[int]) -> str:
    values = sorted(week for week in weeks if week > 0)
    if not values:
        return ""
    ranges: list[str] = []
    start = previous = values[0]
    for week in values[1:]:
        if week == previous + 1:
            previous = week
            continue
        ranges.append(str(start) if start == previous else f"{start}-{previous}")
        start = previous = week
    ranges.append(str(start) if start == previous else f"{start}-{previous}")
    return f"{','.join(ranges)}周"


def export_schedule_workbook(
    destination: str | Path,
    selected: Iterable[Course],
    personal: PersonalSchedule | None,
) -> int:
    """Export the combined timetable using the official personal-schedule layout."""
    period_groups = (
        ("第一大节\n01、02\n08:00~09:40", frozenset({1, 2})),
        ("第二大节\n03、04\n10:00~11:40", frozenset({3, 4})),
        ("第三大节\n05、06\n14:30~16:00", frozenset({5, 6})),
        ("第四大节\n07、08\n16:10~17:40", frozenset({7, 8})),
        ("第五大节\n09、10、11\n19:00~21:35", frozenset({9, 10, 11})),
    )
    display_days = (7, 1, 2, 3, 4, 5, 6)
    day_labels = ("星期日", "星期一", "星期二", "星期三", "星期四", "星期五", "星期六")
    cells: dict[tuple[int, int], list[str]] = {}
    course_names: set[str] = set()

    def add_entry(name: str, teacher: str, classroom: str, slot: TimeSlot) -> None:
        details = f"{name}\n{teacher};{_format_weeks(slot.weeks)}{classroom}\nnull"
        for group_index, (_label, periods) in enumerate(period_groups):
            if slot.periods & periods:
                entries = cells.setdefault((group_index, slot.weekday), [])
                if details not in entries:
                    entries.append(details)
        course_names.add("".join(name.split()).casefold())

    remark_entries: list[str] = []
    if personal:
        for course in personal.courses:
            if course.slots:
                for slot in course.slots:
                    add_entry(course.name, course.teacher, course.classroom, slot)
            else:
                remark_entries.append(course.details.strip() or f"{course.name} {course.teacher}")
                course_names.add("".join(course.name.split()).casefold())
    for course in selected:
        for slot in course.slots:
            add_entry(course.name, course.teacher, course.classroom, slot)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "个人课表"
    sheet.merge_cells("A1:H1")
    sheet["A1"] = "湖南大学 个人课表"
    sheet["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.merge_cells("A2:H2")
    term = personal.term if personal else ""
    sheet["A2"] = f"学年学期：{term or '未识别'}        导出日期：{date.today().isoformat()}"
    sheet["A2"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[2].height = 22
    sheet.append(["节次", *day_labels])

    thin = Side(style="thin", color="FFD8DEE9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", fgColor="FFEAF0F8")
    for column in range(1, 9):
        cell = sheet.cell(3, column)
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[3].height = 24

    for group_index, (label, _periods) in enumerate(period_groups):
        row = group_index + 4
        sheet.cell(row, 1, label)
        for column, day in enumerate(display_days, 2):
            sheet.cell(row, column, "\n\n".join(cells.get((group_index, day), [])))
        for column in range(1, 9):
            cell = sheet.cell(row, column)
            cell.font = Font(name="Microsoft YaHei", size=9)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.row_dimensions[row].height = 72

    note_row = len(period_groups) + 4
    sheet.merge_cells(start_row=note_row, end_row=note_row, start_column=2, end_column=8)
    note_text = ";".join(entry.rstrip(";；") for entry in remark_entries)
    sheet.cell(note_row, 2, f"备注:{note_text}{';' if note_text else ''}")
    for column in range(1, 9):
        cell = sheet.cell(note_row, column)
        cell.font = Font(name="Microsoft YaHei", size=9)
        cell.border = border
        cell.alignment = Alignment(horizontal="left" if column == 2 else "center", vertical="center", wrap_text=True)
    sheet.row_dimensions[note_row].height = max(24, 18 * max(1, len(remark_entries)))
    sheet.column_dimensions["A"].width = 20
    for column in "BCDEFGH":
        sheet.column_dimensions[column].width = 26
    sheet.freeze_panes = "B4"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    try:
        workbook.save(destination)
    except Exception as exc:
        raise DataFormatError(f"无法保存课表：{exc}") from exc
    return len(course_names)


def course_key(course: Course) -> str:
    return "|".join((course.course_id, course.notice_id, course.name, course.teacher, course.raw_time))


def save_plan(path: str | Path, *, catalog: str, selected: Iterable[Course], personal: PersonalSchedule | None, week: int, visible_columns: dict[str, bool], field_filters: dict[str, str], quick_filters: dict[str, str], splitter_sizes: list[int]) -> None:
    payload = {
        "version": 1,
        "catalog": catalog,
        "selected_course_keys": [course_key(course) for course in selected],
        "personal_schedule": personal.to_dict() if personal else None,
        "week": week,
        "visible_columns": visible_columns,
        "field_filters": field_filters,
        "quick_filters": quick_filters,
        "splitter_sizes": splitter_sizes,
    }
    Path(path).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_plan(path: str | Path) -> dict[str, Any]:
    try:
        payload = json.loads(Path(path).read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise DataFormatError(f"无法读取选课方案：{exc}") from exc
    if payload.get("version") != 1 or not isinstance(payload.get("selected_course_keys"), list):
        raise DataFormatError("不支持的选课方案格式。")
    return payload
